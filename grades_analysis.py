import marimo

__generated_with = "0.16.3"
app = marimo.App(width="medium")


@app.cell(hide_code=True)
def _():
    import marimo as mo
    return (mo,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    # Student Grades Analysis

    This notebook demonstrates how to work with Excel data in Python using pandas.
    We'll explore:
    - **Horizontal calculations**: Computing values across columns (e.g., average grade per student)
    - **Vertical calculations**: Computing values down rows (e.g., average per subject)
    - **Pivot tables**: Reshaping and summarizing data
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""## 1. Loading the Data""")
    return


@app.cell
def _():
    import pandas as pd
    import numpy as np
    import openpyxl

    # Load the Excel file
    df = pd.read_excel('data/grades.xlsx')

    # Create a Student name column by combining first and last names
    df['Student'] = df['first_name'] + ' ' + df['last_name']

    df
    return df, pd


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    ## 2. Horizontal Calculations

    Horizontal calculations work **across columns** for each row.
    This is useful when you want to aggregate data for individual records (e.g., calculating total or average scores for each student).
    """
    )
    return


@app.cell
def _(df):
    # Calculate the average grade for each student across all grade columns
    # We select only the numeric grade columns
    subject_columns = ['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4', 'essay_1', 'essay_2', 'essay_3', 'midterm_exam', 'final_exam']

    # Create a copy with the average grade column
    df_with_avg = df.copy()
    df_with_avg['Average'] = df_with_avg[subject_columns].mean(axis=1)  # axis=1 means across columns
    df_with_avg['Total'] = df_with_avg[subject_columns].sum(axis=1)

    # Round to 2 decimal places for better display
    df_with_avg['Average'] = df_with_avg['Average'].round(2)

    df_with_avg
    return (subject_columns,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **Key concept**: `axis=1` tells pandas to perform the operation across columns (horizontally).

    - `mean(axis=1)`: Calculate mean across columns for each row
    - `sum(axis=1)`: Calculate sum across columns for each row
    - `max(axis=1)`: Find maximum value across columns for each row
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### More Horizontal Calculations""")
    return


@app.cell
def _(df, subject_columns):
    # Additional horizontal calculations
    horizontal_stats = df.copy()
    horizontal_stats['Average'] = df[subject_columns].mean(axis=1).round(2)
    horizontal_stats['Min_Grade'] = df[subject_columns].min(axis=1)
    horizontal_stats['Max_Grade'] = df[subject_columns].max(axis=1)
    horizontal_stats['Range'] = horizontal_stats['Max_Grade'] - horizontal_stats['Min_Grade']

    horizontal_stats
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Averages by Assignment Type""")
    return


@app.cell
def _(df):
    # Calculate average for each student by assignment type
    df_by_type = df.copy()

    # Quiz average
    df_by_type['Quiz_Average'] = df_by_type[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].mean(axis=1).round(2)

    # Essay average
    df_by_type['Essay_Average'] = df_by_type[['essay_1', 'essay_2', 'essay_3']].mean(axis=1).round(2)

    # Exam average
    df_by_type['Exam_Average'] = df_by_type[['midterm_exam', 'final_exam']].mean(axis=1).round(2)

    # Select relevant columns for display
    df_by_type[['Student', 'Quiz_Average', 'Essay_Average', 'Exam_Average']]
    return (df_by_type,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This table shows each student's average performance by assignment type:
    - **Quiz_Average**: Average across all 4 quizzes
    - **Essay_Average**: Average across all 3 essays
    - **Exam_Average**: Average of midterm and final exams

    This helps identify whether students perform differently on different types of assessments.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    ## 3. Vertical Calculations

    Vertical calculations work **down rows** for each column.
    This is useful when you want to aggregate data across all records (e.g., calculating class average per subject).
    """
    )
    return


@app.cell
def _(df, subject_columns):
    # Calculate statistics for each subject across all students
    vertical_stats = df[subject_columns].agg(['mean', 'median', 'std', 'min', 'max']).round(2)
    vertical_stats
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        r"""
    **Key concept**: `axis=0` (default) tells pandas to perform the operation down rows (vertically).

    - `mean(axis=0)` or `mean()`: Calculate mean down rows for each column
    - `sum()`: Calculate sum down rows for each column
    - `agg([...])`: Apply multiple aggregation functions at once
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Class Performance Summary""")
    return


@app.cell
def _(df, pd, subject_columns):
    # Create a summary showing vertical calculations
    summary_data = {
        'Subject': subject_columns,
        'Class Average': [df[col].mean().round(2) for col in subject_columns],
        'Highest Score': [df[col].max() for col in subject_columns],
        'Lowest Score': [df[col].min() for col in subject_columns],
        'Std Deviation': [df[col].std().round(2) for col in subject_columns]
    }

    class_summary = pd.DataFrame(summary_data)
    class_summary
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Class Averages by Assignment Type""")
    return


@app.cell
def _(df, pd):
    # Calculate overall class performance by assignment type
    type_summary = pd.DataFrame({
        'Assignment Type': ['Quizzes', 'Essays', 'Exams'],
        'Class Average': [
            df[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].values.mean().round(2),
            df[['essay_1', 'essay_2', 'essay_3']].values.mean().round(2),
            df[['midterm_exam', 'final_exam']].values.mean().round(2)
        ],
        'Highest Score': [
            df[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].values.max(),
            df[['essay_1', 'essay_2', 'essay_3']].values.max(),
            df[['midterm_exam', 'final_exam']].values.max()
        ],
        'Lowest Score': [
            df[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].values.min(),
            df[['essay_1', 'essay_2', 'essay_3']].values.min(),
            df[['midterm_exam', 'final_exam']].values.min()
        ],
        'Std Deviation': [
            df[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].values.std().round(2),
            df[['essay_1', 'essay_2', 'essay_3']].values.std().round(2),
            df[['midterm_exam', 'final_exam']].values.std().round(2)
        ]
    })

    type_summary
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This summary shows class-wide performance across different assignment types.
    You can quickly compare how the class performs on quizzes vs essays vs exams.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    ## 4. Pivot Tables

    Pivot tables are powerful tools for reshaping and summarizing data. They allow you to:
    - Group data by categories
    - Apply aggregation functions
    - Create cross-tabulations
    - Analyze data from different perspectives
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Example 1: Grade Statistics Summary""")
    return


@app.cell
def _(df, pd, subject_columns):
    # Create a summary statistics table for all grade columns
    pivot1 = pd.DataFrame({
        'Assignment': subject_columns,
        'Mean': [df[col].mean().round(2) for col in subject_columns],
        'Median': [df[col].median() for col in subject_columns],
        'Std': [df[col].std().round(2) for col in subject_columns]
    })

    pivot1
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This summary table shows statistics for each assignment across all students.
    We can see the mean, median, and standard deviation for each grade component.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Example 2: Quiz vs Essay vs Exam Performance""")
    return


@app.cell
def _(df, pd):
    # Compare performance across different assignment types
    assignment_types = pd.DataFrame({
        'Type': ['Quizzes', 'Essays', 'Exams'],
        'Average': [
            df[['quiz_1', 'quiz_2', 'quiz_3', 'quiz_4']].mean().mean().round(2),
            df[['essay_1', 'essay_2', 'essay_3']].mean().mean().round(2),
            df[['midterm_exam', 'final_exam']].mean().mean().round(2)
        ]
    })

    assignment_types
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This pivot table shows multiple statistics for Math scores by class:
    - **mean**: Average Math score per class
    - **min**: Lowest Math score per class
    - **max**: Highest Math score per class
    - **count**: Number of students per class
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Example 3: Custom Performance Categories""")
    return


@app.cell
def _(df, subject_columns):
    # Create a performance category based on average grade
    df_enhanced = df.copy()
    df_enhanced['Average'] = df_enhanced[subject_columns].mean(axis=1)

    # Categorize students
    def categorize_performance(avg):
        if avg >= 90:
            return 'Excellent'
        elif avg >= 80:
            return 'Good'
        elif avg >= 70:
            return 'Satisfactory'
        else:
            return 'Needs Improvement'

    df_enhanced['Performance'] = df_enhanced['Average'].apply(categorize_performance)

    df_enhanced
    return (df_enhanced,)


@app.cell
def _(df_enhanced, pd):
    # Count students in each performance category
    pivot3 = df_enhanced['Performance'].value_counts().sort_index()

    # Convert to DataFrame for better display
    performance_summary = pd.DataFrame({
        'Performance Category': pivot3.index,
        'Number of Students': pivot3.values
    })

    performance_summary
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This summary shows how many students fall into each performance category based on their average grade.
    This helps understand the overall distribution of student performance.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""### Example 4: Advanced Pivot - Assignment Averages by Performance Category""")
    return


@app.cell
def _(df_enhanced, pd, subject_columns):
    # Reshape data to long format for more complex pivoting
    df_long = pd.melt(
        df_enhanced,
        id_vars=['Student', 'Performance'],
        value_vars=subject_columns,
        var_name='Assignment',
        value_name='Score'
    )

    # Pivot: Average score by Performance category and Assignment
    pivot4 = pd.pivot_table(
        df_long,
        values='Score',
        index='Performance',
        columns='Assignment',
        aggfunc='mean'
    ).round(2)

    pivot4
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    This pivot table shows the average score for each assignment, broken down by performance category.
    It helps identify if certain performance categories struggle with specific types of assignments (quizzes, essays, or exams).

    **Note**: We used `pd.melt()` to transform the data from wide to long format, making it easier to create this type of pivot.
    """
    )
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(
        """
    ## Summary

    ### Key Takeaways:

    **Horizontal Calculations (axis=1)**:
    - Work across columns for each row
    - Example: `df.mean(axis=1)` calculates row-wise means
    - Use for per-record aggregations

    **Vertical Calculations (axis=0 or default)**:
    - Work down rows for each column
    - Example: `df.mean()` calculates column-wise means
    - Use for overall statistics per feature

    **Pivot Tables**:
    - `pd.pivot_table()` creates summary tables
    - Parameters:
      - `values`: columns to aggregate
      - `index`: rows of pivot table
      - `columns`: columns of pivot table
      - `aggfunc`: aggregation function(s)
    - Use `pd.melt()` to reshape data when needed
    """
    )
    return


if __name__ == "__main__":
    app.run()
